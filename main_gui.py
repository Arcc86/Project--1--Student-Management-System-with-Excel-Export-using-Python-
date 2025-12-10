import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from classs import student, book, reunion, admin
from openpyxl import Workbook

# Global List for save the students
students = []

class StudentManagerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Gestión de Estudiantes")
        self.root.geometry("700x600")
        self.root.iconbitmap('excel_icon.ico')
        
        # Variables para los campos de entrada
        self.name_var = tk.StringVar()
        self.email_var = tk.StringVar()
        self.phone_var = tk.StringVar()
        self.account_var = tk.StringVar()
        self.age_var = tk.StringVar()
        self.books_var = tk.StringVar()
        
        self.create_widgets()
    
    def create_widgets(self):
        # Title
        title_label = tk.Label(self.root, text="Sistema de Gestión de Estudiantes", 
                              font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        
        input_frame = tk.Frame(self.root)
        input_frame.pack(pady=10, padx=20, fill="x")
        
       
        tk.Label(input_frame, text="Nombre:").grid(row=0, column=0, sticky="w", pady=2)
        tk.Entry(input_frame, textvariable=self.name_var, width=30).grid(row=0, column=1, pady=2, padx=(10,0))
        
        tk.Label(input_frame, text="Correo:").grid(row=1, column=0, sticky="w", pady=2)
        tk.Entry(input_frame, textvariable=self.email_var, width=30).grid(row=1, column=1, pady=2, padx=(10,0))
        
        tk.Label(input_frame, text="Teléfono:").grid(row=2, column=0, sticky="w", pady=2)
        tk.Entry(input_frame, textvariable=self.phone_var, width=30).grid(row=2, column=1, pady=2, padx=(10,0))
        
        tk.Label(input_frame, text="Cuenta:").grid(row=3, column=0, sticky="w", pady=2)
        tk.Entry(input_frame, textvariable=self.account_var, width=30).grid(row=3, column=1, pady=2, padx=(10,0))
        
        tk.Label(input_frame, text="Edad:").grid(row=4, column=0, sticky="w", pady=2)
        tk.Entry(input_frame, textvariable=self.age_var, width=30).grid(row=4, column=1, pady=2, padx=(10,0))
        
        # Buttons
        button_frame = tk.Frame(self.root)
        button_frame.pack(pady=10)
        
        tk.Button(button_frame, text="Añadir Estudiante", 
                 command=self.add_student, bg="green", fg="white").pack(side="left", padx=5)
        
        tk.Button(button_frame, text="Guardar en Excel", 
                 command=self.save_to_excel, bg="blue", fg="white").pack(side="left", padx=5)
        
        tk.Button(button_frame, text="Mostrar Estudiantes", 
                 command=self.show_students, bg="orange", fg="white").pack(side="left", padx=5)
        
        # Students list
        self.listbox = tk.Listbox(self.root, width=80, height=15)
        self.listbox.pack(pady=10, padx=20, fill="both", expand=True)
    
    def add_student(self):
        # Validate
        if not all([self.name_var.get(), self.email_var.get(), self.phone_var.get()]):
            messagebox.showerror("Error", "Por favor, completa todos los campos obligatorios")
            return
        
        
        nuevo_estudiante = {
            "Nombre": self.name_var.get(),
            "Email": self.email_var.get(),
            "Telefono": self.phone_var.get(),
            "Cuenta": self.account_var.get(),
            "Edad": self.age_var.get(),
        }
        
        # Add to list
        students.append(nuevo_estudiante)
        
        # Clear_Fields
        self.clear_fields()
        
        # Show Msg
        messagebox.showinfo("Éxito", f"Estudiante {nuevo_estudiante['Nombre']} añadido")
    
    def clear_fields(self):
        self.name_var.set("")
        self.email_var.set("")
        self.phone_var.set("")
        self.account_var.set("")
        self.age_var.set("")
        self.books_var.set("")
    
    def save_to_excel(self):
        if not students:
            messagebox.showwarning("Advertencia", "No hay estudiantes para guardar")
            return
        
        # Ask where do you want to save the file.
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
        )
        
        if filename:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Gestion de estudiantes"
                
                
                headers = ["Nombre", "Email", "Telefono", "Cuenta", "Edad"]
                ws.append(headers)
                
                
                for student in students:
                    row = [student[header] for header in headers]
                    ws.append(row)
                
                wb.save(filename)
                messagebox.showinfo("Éxito", f"Datos guardados en {filename}")
            except Exception as e:
                messagebox.showerror("Error", f"Error al guardar el archivo: {str(e)}")
    
    def show_students(self):
        
        self.listbox.delete(0, tk.END)
        
        if not students:
            self.listbox.insert(tk.END, "No hay estudiantes registrados")
            return
        
        
        for i, student in enumerate(students, 1):
            student_info = f"{i}. {student['Nombre']} - {student['Email']} - {student['Telefono']}"
            self.listbox.insert(tk.END, student_info)

def main():
    root = tk.Tk()
    app = StudentManagerGUI(root)
    root.mainloop()

def save_to_excel(self):
    if not students:
        messagebox.showwarning("Advertencia", "No hay estudiantes para guardar")
        return
    
    filename = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
    )
    
    if filename:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Estudiantes"
            
            # Colores por encabezado (cada columna tendrá su color)
            header_colors = {
                "Nombre": "4472C4",    # Azul
                "Email": "70AD47",     # Verde
                "Telefono": "FFC000",  # Amarillo
                "Cuenta": "ED7D31",    # Naranja
                "Edad": "5B9BD5",      # Azul claro
                "Libros leidos": "9E480E"  # Marrón
            }
            
            # Encabezados
            headers = ["Nombre", "Email", "Telefono", "Cuenta", "Edad", "Libros leidos"]
            ws.append(headers)
            
            # Formatear encabezados con colores individuales
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                
                # Color de fondo para cada encabezado
                color_code = header_colors[header]
                header_fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
                
                # Estilo del texto
                header_font = Font(bold=True, color="FFFFFF")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                
                # Bordes para encabezados
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
            
            # Añadir datos
            for student in students:
                row_data = [student[header] for header in headers]
                ws.append(row_data)
                
                # Formatear la fila con bordes
                row_num = ws.max_row
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = Border(
                        left=Side(style='thin', color='CCCCCC'),
                        right=Side(style='thin', color='CCCCCC'),
                        top=Side(style='thin', color='CCCCCC'),
                        bottom=Side(style='thin', color='CCCCCC')
                    )
            
            # Auto-ajustar columnas
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 30)  # Máximo 30 caracteres
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Crear tabla con estilo
            tab = Table(displayName="TablaEstudiantes", ref=f"A1:F{len(students) + 1}")
            
            # Estilo de la tabla
            style = TableStyleInfo(
                name="TableStyleMedium9", 
                showFirstColumn=False,
                showLastColumn=False, 
                showRowStripes=True, 
                showColumnStripes=True
            )
            tab.tableStyleInfo = style
            
            # Añadir la tabla a la hoja
            ws.add_table(tab)
            
            wb.save(filename)
            messagebox.showinfo("Éxito", f"Datos guardados en {filename} con tabla y colores")
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar el archivo: {str(e)}")
            
if __name__ == '__main__':
    main()