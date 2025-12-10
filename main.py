from classs import student, book, reunion, admin
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def load_data():
    # Load data from file or initialize empty data
    return {"Students": [], "Meetings": []}

def show_menu():
    print("\n1. Add student")
    print("2. List students")
    print("3. Export to Excel")
    print("4. Exit")

def export_to_excel(students):
    # Export the list of students to an Excel file with format
    if not students:
        print("No students to export.")
        return
    
    filename = input("Excel file name (without extension): ") + ".xlsx"
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Colors by header
        header_colors = {
            "Name": "4472C4",    # Blue
            "Email": "70AD47",     # Green
            "Phone Number": "FFC000",  # Yellow
            "Account": "ED7D31",    # Orange
            "Age": "5B9BD5",      # Light blue
            "Books read": "9E480E"  # Brown
        }
        
        # Headers
        headers = ["Name", "Email", "Phone Number", "Account", "Age", "Books read"]
        ws.append(headers)
        
        # Format headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            
            color_code = header_colors[header]
            header_fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
        
        # Add data
        for student in students:
            row_data = [student[header] for header in headers]
            ws.append(row_data)
            
            # Format row with borders
            row_num = ws.max_row
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
        
        # Auto-adjust columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15  # Width of 15 for each column
        
        wb.save(filename)
        print(f"Data exported to {filename} with professional format")
    except Exception as e:
        print(f"Error exporting: {str(e)}")

def main():
    # Load data
    data = load_data()
    students = data.get("Students", [])  # students list
    reunions = data.get("Meetings", [])
    
    print(f'\n =============================\n '
          f'=========================================\n'
          f'Student Management System \n ===========================\n'
          f'=========================================')

    while True:
        show_menu()
        option = input("Choose an option: ")

        if option == "1":
            # Add student
            name = input("Name: ")
            email = input("Email: ")
            phone = input("Phone: ")
            account = input("Account number: ")
            age = input("Age: ")
            books_read = input("Number of books read: ")

            # Create dictionary with data
            new_student = {
                "Name": name,
                "Email": email,
                "Phone Number": phone,
                "Account": account,
                "Age": age,
                "Books Read": books_read
            }

            # Add to list
            students.append(new_student)
            print(f"Student {name} added.")

        elif option == "2":
            # List students
            print("\n --- Registered Students --- ")
            if not students:  # students is defined here
                print("No students registered yet...")
            else:
                for e in students:
                    print(f"- {e['Name']} - {e['Email']} - {e['Phone Number']} - {e['Account']} - {e['Age']} - {e['Books Read']}")

        elif option == "3":
            # Export to Excel
            export_to_excel(students)

        elif option == "4":
            print("Exiting...")
            break

if __name__ == '__main__':
    main()